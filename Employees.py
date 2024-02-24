# code to get birth year and age from DoB

import openpyxl as xl
from openpyxl.utils import get_column_letter
from datetime import datetime


wb = xl.load_workbook("Employees.xlsx")
sheet = wb["Sheet1"]

# code to get year of birth from DoB
sheet.cell(row=2, column=6, value='Year of Birth')

# Iterate over the rows starting from the second row
for row in range(3, sheet.max_row + 1):
    # Get the date from the 5th column
    date_cell = sheet.cell(row=row, column=5)
    date_value = date_cell.value

    # Convert the date to a datetime object
    if isinstance(date_value, datetime):
        year = date_value.year
        # Add the year to the new column
        sheet.cell(row=row, column=6, value=year)

# Code to get age from DoB
sheet.cell(row=2, column=7, value='Age')

# Iterate over the rows starting from the second row
for row in range(3, sheet.max_row + 1):
    # Get the date from the 5th column
    date_cell = sheet.cell(row=row, column=5)
    date_value = date_cell.value

    # Convert the date to a datetime object
    if isinstance(date_value, datetime):
        today = datetime.today()
        age = today.year - date_value.year - ((today.month, today.day) < (date_value.month, date_value.day))
        # Add the age to the new column
        sheet.cell(row=row, column=7, value= age)

wb.save('EmployeesAge.xlsx')






